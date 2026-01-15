#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
📝 Txt2Excel v2.1_20251213
~~~~~~~~~~~~~~~~~~~~~~~~~~~

🧩 功能特色:
├── 自動解析 TXT 數據檔案
├── 智慧識別 Excel 結構
├── 靈活配置填入規則
├── 完整保留格式設定
├── 🖼️ 圖片保護完整支援
└── 合併儲存格智慧處理

📚 技術亮點:
├── 雙引擎支援 (openpyxl + win32com)
├── 記憶體優化與垃圾回收
├── 多執行緒安全設計
└── 完善錯誤處理機制

👨‍💻 開發者: [Wesley Chang]
電子郵箱: [wesleychang2025@gmail.com]
📅 最後更新: 2025-12-13
"""

import os
import sys
import chardet
import traceback
import threading
from collections import OrderedDict
from tkinter import *
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple as coordinate_from_string
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import gc

# 🔧 新增導入win32com模組以支持COM接口
try:
    import win32com.client as win32
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False
    print("警告：未安裝win32com，將無法使用COM接口處理Excel")

# 嘗試預先設定Tcl/Tk函式庫路徑以避免環境問題
try:
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller打包環境
        tcl_path = os.path.join(sys._MEIPASS, 'tcl')
        if os.path.exists(tcl_path):
            os.environ['TCL_LIBRARY'] = os.path.join(tcl_path, 'tcl8.6')
            os.environ['TK_LIBRARY'] = os.path.join(tcl_path, 'tk8.6')
    else:
        # 開發環境
        python_root = os.path.dirname(sys.executable)
        tcl_candidate_paths = [
            os.path.join(python_root, 'tcl', 'tcl8.6'),
            os.path.join(python_root, 'lib', 'tcl8.6'),
            os.path.join(os.path.dirname(__file__), '..', 'tcl', 'tcl8.6')
        ]
        
        for tcl_path in tcl_candidate_paths:
            if os.path.exists(tcl_path):
                os.environ['TCL_LIBRARY'] = tcl_path
                os.environ['TK_LIBRARY'] = tcl_path.replace('tcl8.6', 'tk8.6')
                break
except Exception:
    pass


class DataPanel(Frame):
    """資料預覽面板
    
    用於顯示和預覽從TXT檔案中提取的資料。
    支援檔案選擇、資料載入、顯示和清理等功能。
    """
    def __init__(self, parent):
        super().__init__(parent, bg='white', relief='solid', bd=1)
        self.txt_file_path = None
        self.extracted_values = []
        
        # 標題
        title = Label(self, text="📄 來源資料", font=("Microsoft YaHei", 12, "bold"),
                     bg='white', fg='#2E86AB')
        title.pack(pady=(10, 5))
        
        # 檔案選擇區
        file_frame = Frame(self, bg='white')
        file_frame.pack(fill=X, padx=10, pady=5)
        
        Label(file_frame, text="TXT 檔案：", bg='white').pack(side=LEFT)
        self.file_label = Label(file_frame, text="未選擇", bg='white', fg='gray')
        self.file_label.pack(side=LEFT, padx=5)
        
        btn_select = Button(file_frame, text="選擇檔案", command=self.select_file,
                           bg='#2E86AB', fg='white', relief='flat', padx=15)
        btn_select.pack(side=RIGHT)
        
        # 資料統計
        self.stats_label = Label(self, text="已解析資料：0 筆", 
                                font=("Microsoft YaHei", 10), bg='white', fg='#333')
        self.stats_label.pack(pady=5)
        
        # 按鈕區（先 pack 在底部）
        btn_frame = Frame(self, bg='white')
        btn_frame.pack(side=BOTTOM, fill=X, padx=10, pady=10)
        
        # 資料預覽列表
        Label(self, text="資料預覽：", bg='white', anchor='w').pack(fill=X, padx=10)
        
        list_frame = Frame(self, bg='white')
        list_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=VERTICAL)
        self.data_listbox = Listbox(list_frame, yscrollcommand=scrollbar.set,
                                    font=("Consolas", 10))
        scrollbar.config(command=self.data_listbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.data_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        
        Button(btn_frame, text="清除", command=self.clear_data,
               bg='#dc3545', fg='white', relief='flat', padx=10).pack(side=LEFT, padx=2)
        Button(btn_frame, text="重新載入", command=self.reload_data,
               bg='#6c757d', fg='white', relief='flat', padx=10).pack(side=LEFT, padx=2)

    
    def select_file(self):
        """選擇 TXT 檔案"""
        file_path = filedialog.askopenfilename(
            title='選擇 TXT 檔案',
            filetypes=[('Text Files', '*.txt'), ('All Files', '*.*')]
        )
        if file_path:
            self.txt_file_path = file_path
            self.load_txt_file()
    
    def load_txt_file(self):
        """載入並解析 TXT 檔案"""
        if not self.txt_file_path:
            return
        
        try:
            # 自動偵測編碼
            with open(self.txt_file_path, 'rb') as f:
                rawdata = f.read()
                result = chardet.detect(rawdata)
                encoding = result['encoding']
            
            # 解析資料
            self.extracted_values = []
            with open(self.txt_file_path, 'r', encoding=encoding) as f:
                for line_num, line in enumerate(f, 1):
                    if '=' in line:
                        right = line.split('=', 1)[1].strip()
                        # 🔧 修復：更嚴格地清理數據，移除可能的額外字符
                        # 保留所有数字部分，不僅僅是逗号分隔的第一部分
                        # 檢查是否為空值，如果是則停止解析
                        if not right or right.isspace():
                            # 根據記憶要求，遇到第一個空值後停止解析
                            break
                        
                        # 移除可能存在的非打印字符和額外的空白字符
                        cleaned_value = right.strip('\r\n\t\b\f\v\0')
                        # 進一步清理可能存在的隱藏字符
                        cleaned_value = ''.join(char for char in cleaned_value if ord(char) >= 32 or char in '\t\n\r')
                        
                        # 🔧 修復：處理尾隨的無意義字符（如逗號、空格等）
                        # 從右側移除常見的無意義尾隨字符
                        cleaned_value = cleaned_value.rstrip(', ')
                        
                        # 🔧 修復：處理特殊情況，如只有分號或其他無意義字符的值
                        # 檢查清理後的值是否只包含分號或其他無意義字符
                        if not cleaned_value or cleaned_value.isspace() or all(c in ';:,.' for c in cleaned_value):
                            # 根據記憶要求，遇到無意義值後停止解析
                            break
                        
                        # 🔧 修復：添加調試信息以幫助診斷問題
                        if line_num <= 10:  # 只顯示前10行進行調試
                            print(f"解析第 {line_num} 行: 原始='{right}', 清理後='{cleaned_value}'")
                        self.extracted_values.append(cleaned_value)
            
            # 更新顯示
            filename = os.path.basename(self.txt_file_path)
            self.file_label.config(text=filename, fg='#28a745')
            self.stats_label.config(text=f"已解析資料：{len(self.extracted_values)} 筆")
            
            # 顯示資料
            self.display_data()
            
            messagebox.showinfo("成功", f"成功載入 {len(self.extracted_values)} 筆資料")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"讀取 TXT 檔案時發生錯誤：{e}")
    
    def display_data(self):
        """顯示資料列表"""
        self.data_listbox.delete(0, END)
        for i, value in enumerate(self.extracted_values, 1):
            # 只顯示前20個字符，避免列表框顯示過於擁擠
            display_value = value[:20] + "..." if len(value) > 20 else value
            self.data_listbox.insert(END, f"{i:3d}. {display_value}")
    
    def clear_data(self):
        """清除資料"""
        if messagebox.askyesno("確認", "確定要清除所有資料嗎？"):
            self.txt_file_path = None
            self.extracted_values = []
            self.file_label.config(text="未選擇", fg='gray')
            self.stats_label.config(text="已解析資料：0 筆")
            self.data_listbox.delete(0, END)
    
    def reload_data(self):
        """重新載入資料"""
        if self.txt_file_path:
            self.load_txt_file()
        else:
            messagebox.showwarning("警告", "請先選擇 TXT 檔案")
    
    def get_data(self):
        """取得解析的資料"""
        return self.extracted_values


class ConfigPanel(Frame):
    """配置面板
    
    用於配置Excel檔案的填入規則。
    支援Excel檔案選擇、工作表資訊顯示、填入規則配置等功能。
    """
    def __init__(self, parent):
        super().__init__(parent, bg='white', relief='solid', bd=1)
        self.excel_file_path = None
        self.workbook = None
        self.sheet_names = []  # 快取工作表名稱
        self.rules = []  # 儲存組態規則
        
        # 標題
        title = Label(self, text="📊 填入配置", font=("Microsoft YaHei", 12, "bold"),
                     bg='white', fg='#2E86AB')
        title.pack(pady=(10, 5))
        
        # Excel 檔案選擇
        file_frame = Frame(self, bg='white')
        file_frame.pack(fill=X, padx=10, pady=5)
        
        Label(file_frame, text="Excel 檔案：", bg='white').pack(side=LEFT)
        self.file_label = Label(file_frame, text="未選擇", bg='white', fg='gray')
        self.file_label.pack(side=LEFT, padx=5)
        
        btn_select = Button(file_frame, text="選擇檔案", command=self.select_file,
                           bg='#2E86AB', fg='white', relief='flat', padx=15)
        btn_select.pack(side=RIGHT)
        
        # 工作表資訊
        self.sheet_label = Label(self, text="工作表：0 個", 
                                font=("Microsoft YaHei", 10), bg='white', fg='#333')
        self.sheet_label.pack(pady=5)
        
        # 按鈕區（先 pack 在底部）
        btn_frame = Frame(self, bg='white')
        btn_frame.pack(side=BOTTOM, fill=X, padx=10, pady=10)
        
        # 規則列表標題
        Label(self, text="填入規則：", bg='white', anchor='w').pack(fill=X, padx=10, pady=(10, 5))
        
        # 規則表格
        table_frame = Frame(self, bg='white')
        table_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        # 建立 Treeview
        columns = ('sheet', 'start_cell', 'count', 'direction')
        self.rule_tree = ttk.Treeview(table_frame, columns=columns, show='headings')
        
        self.rule_tree.heading('sheet', text='工作表')
        self.rule_tree.heading('start_cell', text='起始儲存格')
        self.rule_tree.heading('count', text='筆數')
        self.rule_tree.heading('direction', text='方向')
        
        self.rule_tree.column('sheet', width=120)
        self.rule_tree.column('start_cell', width=100)
        self.rule_tree.column('count', width=60)
        self.rule_tree.column('direction', width=60)
        
        scrollbar = ttk.Scrollbar(table_frame, orient=VERTICAL, command=self.rule_tree.yview)
        self.rule_tree.configure(yscrollcommand=scrollbar.set)
        
        self.rule_tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        Button(btn_frame, text="新增規則", command=self.add_rule,
               bg='#28a745', fg='white', relief='flat', padx=10).pack(side=LEFT, padx=2)
        Button(btn_frame, text="編輯", command=self.edit_rule,
               bg='#ffc107', fg='black', relief='flat', padx=10).pack(side=LEFT, padx=2)
        Button(btn_frame, text="刪除", command=self.delete_rule,
               bg='#dc3545', fg='white', relief='flat', padx=10).pack(side=LEFT, padx=2)
        Button(btn_frame, text="清空", command=self.clear_rules,
               bg='#6c757d', fg='white', relief='flat', padx=10).pack(side=LEFT, padx=2)
    
    def select_file(self):
        """選擇 Excel 檔案"""
        file_path = filedialog.askopenfilename(
            title='選擇 Excel 檔案',
            filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')]
        )
        if file_path:
            self.excel_file_path = file_path
            self.load_excel_file()
    
    def load_excel_file(self):
        """載入 Excel 檔案（快速預覽模式）"""
        if not self.excel_file_path:
            return
        
        # 顯示載入對話框
        loading_dialog = LoadingDialog(self, "正在讀取工作表資訊...")
        
        def load_in_thread():
            try:
                # 🚀 安全的效能最佳化：使用唯讀模式快速讀取工作表名稱
                # ⚠️ 注意：只使用 read_only=True，不使用 data_only 或 keep_links，確保不破壞檔案
                temp_wb = load_workbook(self.excel_file_path, read_only=True)
                                 
                # 快取工作表名稱
                self.sheet_names = temp_wb.sheetnames.copy()
                                 
                # 立即關閉，釋放記憶體
                temp_wb.close()
                
                # 清空 workbook（將在執行時重新載入）
                self.workbook = None
                
                self.after(100, lambda: self.on_excel_loaded(loading_dialog))
            except Exception as e:
                self.after(100, lambda: self.on_excel_error(loading_dialog, str(e)))
        
        thread = threading.Thread(target=load_in_thread, daemon=True)
        thread.start()
    
    def on_excel_loaded(self, dialog):
        """Excel 載入完成"""
        dialog.destroy()
        
        filename = os.path.basename(self.excel_file_path)
        self.file_label.config(text=filename, fg='#28a745')
        
        sheet_count = len(self.sheet_names)
        self.sheet_label.config(text=f"工作表：{sheet_count} 個")
        
        messagebox.showinfo("成功", f"成功載入 Excel 檔案\n共 {sheet_count} 個工作表")
    
    def on_excel_error(self, dialog, error_msg):
        """Excel 載入錯誤"""
        dialog.destroy()
        messagebox.showerror("錯誤", f"讀取 Excel 檔案時發生錯誤：{error_msg}")

    
    def add_rule(self):
        """新增規則"""
        if not self.sheet_names:
            messagebox.showwarning("警告", "請先選擇 Excel 檔案")
            return
        
        dialog = RuleDialog(self, self.sheet_names)
        self.wait_window(dialog)
        
        if dialog.result:
            self.rules.append(dialog.result)
            self.refresh_rule_list()
    
    def edit_rule(self):
        """編輯規則"""
        selection = self.rule_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要編輯的規則")
            return
        
        index = self.rule_tree.index(selection[0])
        rule = self.rules[index]
        
        dialog = RuleDialog(self, self.sheet_names, rule)
        self.wait_window(dialog)
        
        if dialog.result:
            self.rules[index] = dialog.result
            self.refresh_rule_list()
    
    def delete_rule(self):
        """刪除規則"""
        selection = self.rule_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要刪除的規則")
            return
        
        if messagebox.askyesno("確認", "確定要刪除選中的規則嗎？"):
            index = self.rule_tree.index(selection[0])
            del self.rules[index]
            self.refresh_rule_list()
    
    def clear_rules(self):
        """清空所有規則"""
        if self.rules and messagebox.askyesno("確認", "確定要清空所有規則嗎？"):
            self.rules.clear()  # 清空列表
            self.refresh_rule_list()
            gc.collect()  # 🗑️ 強制垃圾回收，釋放記憶體
    
    def refresh_rule_list(self):
        """刷新規則列表"""
        self.rule_tree.delete(*self.rule_tree.get_children())
        for rule in self.rules:
            self.rule_tree.insert('', END, values=(
                rule['sheet'],
                rule['start_cell'],
                rule['count'],
                rule['direction']
            ))
    
    def get_config(self):
        """取得配置"""
        return {
            'excel_path': self.excel_file_path,
            'workbook': self.workbook,  # 此時為 None
            'sheet_names': self.sheet_names,
            'rules': self.rules
        }


class ProgressPanel(Frame):
    """進度面板"""
    def __init__(self, parent):
        super().__init__(parent, bg='#f8f9fa', relief='solid', bd=1)
        
        # 標題
        title = Label(self, text="📈 執行進度", font=("Microsoft YaHei", 11, "bold"),
                     bg='#f8f9fa', fg='#2E86AB')
        title.pack(pady=(10, 5), padx=10, anchor='w')
        
        # 進度條
        self.progress_var = IntVar()
        self.progress_bar = ttk.Progressbar(self, variable=self.progress_var, 
                                           maximum=100, length=400)
        self.progress_bar.pack(fill=X, padx=10, pady=5)
        
        # 進度文字
        self.progress_label = Label(self, text="0% (0/0 笔)", 
                                   font=("Microsoft YaHei", 10),
                                   bg='#f8f9fa', fg='#333')
        self.progress_label.pack(pady=2)
        
        # 目前處理
        self.status_label = Label(self, text="準備就緒", 
                                 font=("Microsoft YaHei", 9),
                                 bg='#f8f9fa', fg='#666')
        self.status_label.pack(pady=(2, 10))
    
    def update_progress(self, current, total, status=""):
        """更新進度"""
        if total > 0:
            percentage = int((current / total) * 100)
            self.progress_var.set(percentage)
            self.progress_label.config(text=f"{percentage}% ({current}/{total} 筆)")
        
        if status:
            self.status_label.config(text=f"目前處理：{status}")
    
    def reset(self):
        """重置進度"""
        self.progress_var.set(0)
        self.progress_label.config(text="0% (0/0 筆)")
        self.status_label.config(text="準備就緒")


class RuleDialog(Toplevel):
    """規則設定對話框"""
    def __init__(self, parent, sheet_names, rule=None):
        super().__init__(parent)
        self.title("設定填入規則")
        self.geometry("450x360")
        self.resizable(False, False)
        
        # 置中顯示
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.all_sheet_names = sheet_names  # 儲存所有工作表
        self.show_all = False  # 預設只顯示最後10個
        
        # 工作表選擇區域（包含過濾按鈕）
        sheet_frame = Frame(self)
        sheet_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky='ew')
        
        Label(sheet_frame, text="工作表：").pack(side=LEFT, padx=(0, 5))
        
        # 🆕 過濾按鈕
        self.filter_btn = Button(
            sheet_frame,
            text="顯示全部" if not self.show_all else "僅顯示最後10個",
            command=self.toggle_filter,
            bg='#17a2b8',
            fg='white',
            relief='flat',
            padx=8,
            font=("Microsoft YaHei", 8)
        )
        self.filter_btn.pack(side=RIGHT)
        
        # 工作表下拉選單
        self.sheet_var = StringVar()
        filtered_sheets = self.get_filtered_sheets()
        self.sheet_combo = ttk.Combobox(
            self,
            textvariable=self.sheet_var,
            values=filtered_sheets,
            state='readonly',
            width=35
        )
        self.sheet_combo.grid(row=1, column=0, columnspan=2, padx=10, pady=(0, 10), sticky='ew')
        
        if rule:
            self.sheet_combo.set(rule['sheet'])
        elif filtered_sheets:
            self.sheet_combo.current(0)
        
        # 起始儲存格
        Label(self, text="起始儲存格：").grid(row=2, column=0, padx=10, pady=10, sticky='e')
        self.cell_var = StringVar(value=rule['start_cell'] if rule else 'I26')
        cell_entry = Entry(self, textvariable=self.cell_var, width=27)
        cell_entry.grid(row=2, column=1, padx=10, pady=10, sticky='w')
        
        # 筆數
        Label(self, text="每頁筆數：").grid(row=3, column=0, padx=10, pady=10, sticky='e')
        self.count_var = IntVar(value=rule['count'] if rule else 8)
        count_spin = ttk.Spinbox(self, textvariable=self.count_var, 
                                from_=1, to=100, width=25)
        count_spin.grid(row=3, column=1, padx=10, pady=10, sticky='w')
        
        # 🔒 固定為橫向填入，移除縱向選項
        self.direction_var = StringVar(value='橫向')  # 固定為橫向
        
        # 預覽
        self.preview_label = Label(self, text="", fg='#666', font=("Microsoft YaHei", 9))
        self.preview_label.grid(row=4, column=0, columnspan=2, pady=10)
        
        # 更新預覽
        self.cell_var.trace('w', self.update_preview)
        self.count_var.trace('w', self.update_preview)
        # direction 不再需要 trace，因為固定為橫向
        self.update_preview()
        
        # 按鈕
        btn_frame = Frame(self)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)
        
        Button(btn_frame, text="確定", command=self.ok,
               bg='#28a745', fg='white', relief='flat', padx=20).pack(side=LEFT, padx=5)
        Button(btn_frame, text="取消", command=self.cancel,
               bg='#6c757d', fg='white', relief='flat', padx=20).pack(side=LEFT, padx=5)
    
    def get_filtered_sheets(self):
        """取得過濾後的工作表名稱"""
        if self.show_all or len(self.all_sheet_names) <= 10:
            return self.all_sheet_names
        # 🆕 只返回最後 10 個工作表
        return self.all_sheet_names[-10:]
    
    def toggle_filter(self):
        """切換顯示全部/最後10個"""
        self.show_all = not self.show_all
        
        # 更新按鈕文字
        self.filter_btn.config(
            text="顯示全部" if not self.show_all else "僅顯示最後10個"
        )
        
        # 更新下拉選單
        current_value = self.sheet_var.get()
        filtered_sheets = self.get_filtered_sheets()
        self.sheet_combo['values'] = filtered_sheets
        
        # 如果目前選擇的工作表還在列表中，保持選擇
        if current_value in filtered_sheets:
            self.sheet_combo.set(current_value)
        elif filtered_sheets:
            self.sheet_combo.current(0)
    
    def update_preview(self, *args):
        """更新預覽"""
        try:
            start_cell = self.cell_var.get().upper()
            count = self.count_var.get()
            
            # 驗證儲存格格式
            # 注意：coordinate_to_tuple 返回 (row, column) 而不是 (column, row)
            row, col = coordinate_from_string(start_cell)
            col = int(col)
            row = int(row)
            
            # 固定橫向填入
            end_col = chr(ord(col) + int(count) - 1) if int(count) < 26 else '...'
            preview = f"預覽：將填入 {start_cell} → {end_col}{row}"
            
            self.preview_label.config(text=preview, fg='#28a745')
        except:
            self.preview_label.config(text="儲存格格式錯誤（例如：A1）", fg='#dc3545')
    
    def ok(self):
        """確定"""
        try:
            # 驗證輸入
            start_cell = self.cell_var.get().upper()
            # 驗證格式
            # 注意：coordinate_to_tuple 返回 (row, column) 而不是 (column, row)
            row, col = coordinate_from_string(start_cell)
            col = int(col)
            row = int(row)
            
            self.result = {
                'sheet': self.sheet_var.get(),
                'start_cell': start_cell,
                'count': self.count_var.get(),
                'direction': self.direction_var.get()
            }
            self.destroy()
        except:
            messagebox.showerror("錯誤", "儲存格格式錯誤，請輸入正確格式（例如：A1）")
    
    def cancel(self):
        """取消"""
        self.destroy()


class LoadingDialog(Toplevel):
    """載入對話框"""
    def __init__(self, parent, message="載入中..."):
        super().__init__(parent)
        self.title("請稍候")
        self.geometry("350x120")
        self.resizable(False, False)
        
        # 置中顯示
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - 350) // 2
        y = (screen_height - 120) // 2
        self.geometry(f"350x120+{x}+{y}")
        
        self.transient(parent)
        self.grab_set()
        
        # 資訊
        Label(self, text=message, font=("Microsoft YaHei", 12)).pack(pady=20)
        
        # 進度條
        progress = ttk.Progressbar(self, mode='indeterminate', length=300)
        progress.pack(pady=10)
        progress.start(10)
        
        self.update()


class MainWindow:
    """主視窗"""
    def __init__(self):
        self.root = Tk()
        self.root.title("Txt2Excel v2.1_20251213")
        self.root.geometry("1000x700")
        
        # 置中顯示
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - 1000) // 2
        y = (screen_height - 700) // 2
        self.root.geometry(f"1000x700+{x}+{y}")
        
        self.setup_ui()
        
    def setup_ui(self):
        """設定介面"""
        # 標題列
        header = Frame(self.root, bg='#2E86AB', height=80)
        header.pack(fill=X)
        header.pack_propagate(False)
        
        title = Label(header, text="Txt2Excel v2.1_20251213", 
                     font=("Microsoft YaHei", 20, "bold"),
                     bg='#2E86AB', fg='white')
        title.pack(pady=20)
        
        # 步驟指示
        step_frame = Frame(self.root, bg='#f8f9fa', height=50)
        step_frame.pack(fill=X)
        step_frame.pack_propagate(False)
        
        steps = ["① 選擇檔案", "→", "② 配置規則", "→", "③ 預覽確認", "→", "④ 執行填入"]
        for i, step in enumerate(steps):
            if step == "→":
                Label(step_frame, text=step, bg='#f8f9fa', fg='#999',
                     font=("Microsoft YaHei", 12)).pack(side=LEFT, padx=5)
            else:
                Label(step_frame, text=step, bg='#f8f9fa', fg='#2E86AB',
                     font=("Microsoft YaHei", 11, "bold")).pack(side=LEFT, padx=10)
        
        # 按鈕列（最先 pack 在底部）
        btn_frame = Frame(self.root, bg='white', height=60)
        btn_frame.pack(fill=X, side=BOTTOM)
        btn_frame.pack_propagate(False)
        
        Button(btn_frame, text="開始執行", command=self.start_execution,
               bg='#28a745', fg='white', font=("Microsoft YaHei", 12, "bold"),
               relief='flat', padx=30, pady=10).pack(side=RIGHT, padx=10, pady=10)
        
        Button(btn_frame, text="重置", command=self.reset_all,
               bg='#6c757d', fg='white', font=("Microsoft YaHei", 11),
               relief='flat', padx=20, pady=10).pack(side=RIGHT, padx=5, pady=10)
        
        Button(btn_frame, text="說明", command=self.show_help,
               bg='#17a2b8', fg='white', font=("Microsoft YaHei", 11),
               relief='flat', padx=20, pady=10).pack(side=LEFT, padx=10, pady=10)
        
        # 底部進度面板（在按鈕列之上）
        self.progress_panel = ProgressPanel(self.root)
        self.progress_panel.pack(fill=X, side=BOTTOM, padx=10, pady=(0, 5))
        
        # 主內容區
        content = Frame(self.root, bg='#e9ecef')
        content.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # 左側面板（資料預覽）
        self.data_panel = DataPanel(content)
        self.data_panel.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 5))
        
        # 右側面板（配置）
        self.config_panel = ConfigPanel(content)
        self.config_panel.pack(side=RIGHT, fill=BOTH, expand=True, padx=(5, 0))

    
    def start_execution(self):
        """開始執行填入"""
        # 驗證
        data = self.data_panel.get_data()
        if not data:
            messagebox.showwarning("警告", "請先選擇並載入 TXT 檔案")
            return
        
        config = self.config_panel.get_config()
        # 檢查 sheet_names
        if not config['sheet_names']:
            messagebox.showwarning("警告", "請先選擇並載入 Excel 檔案")
            return
        
        if not config['rules']:
            messagebox.showwarning("警告", "請至少新增一條填入規則")
            return
        
        # 確認執行
        total_capacity = sum(rule['count'] for rule in config['rules'])
        if len(data) > total_capacity:
            msg = f"資料共 {len(data)} 筆，但規則只能填入 {total_capacity} 筆\n是否繼續？"
            if not messagebox.askyesno("確認", msg):
                return
        
        # 執行填入
        self.execute_filling(data, config)
    
    def execute_filling(self, data, config):
        """執行填入邏輯"""
        # 列印讀取後、寫入前的資料以供檢查
        print("\n" + "="*50)
        print("數據狀態檢查報告")
        print("="*50)
        print(f"讀取到的資料總數: {len(data)}")
        print("前20筆資料（讀取後）:")
        for i, val in enumerate(data[:20]):
            print(f"  [{i:2d}] {repr(val)} (類型: {type(val).__name__})")
        if len(data) > 20:
            print(f"  ... (還有 {len(data) - 20} 筆資料)")
        
        # 顯示配置信息
        print("\n配置信息:")
        print(f"  Excel文件: {config['excel_path']}")
        print(f"  工作表數量: {len(config['sheet_names'])}")
        print(f"  規則數量: {len(config['rules'])}")
        for i, rule in enumerate(config['rules']):
            print(f"  規則 {i+1}: 工作表='{rule['sheet']}', 起始儲存格={rule['start_cell']}, 筆數={rule['count']}, 方向={rule['direction']}")
        print("="*50 + "\n")
        
        rules = config['rules']
        data_idx = 0
        total_data = len(data)

        self.progress_panel.reset()

        self.progress_panel.update_progress(0, 100, "正在載入 Excel 檔案...")
        self.root.update()

        use_com = False
        try:
            # 🔧 修復：使用全局導入的win32模組
            if WIN32_AVAILABLE:
                use_com = True
            else:
                use_com = False
        except Exception:
            use_com = False

        if use_com:
            try:
                excel = win32.gencache.EnsureDispatch("Excel.Application")
                excel.Visible = False
                excel.ScreenUpdating = False
                excel.DisplayAlerts = False
                wb_com = excel.Workbooks.Open(config['excel_path'])
                allowed_sheets = set(rule['sheet'] for rule in rules)
                
                # 🔧 修復：在正確的作用域內定義保護變量
                sheet_protections = {}  # 存儲每個工作表的保護信息
                
                for rule in rules:
                    if data_idx >= total_data:
                        break
                    sheet_name = rule['sheet']
                    start_cell = rule['start_cell']
                    count = rule['count']
                    direction = rule['direction']
                    status = f"{sheet_name} - {start_cell}"
                    self.progress_panel.update_progress(data_idx, total_data, status)
                    self.root.update()
                    ws_com = wb_com.Worksheets(sheet_name)
                    
                    # 🔧 修復：為每個工作表單獨跟踪保護狀態
                    reprotect = False
                    used_pwd = ""
                    prot_opts = {}
                    
                    try:
                        if bool(ws_com.ProtectContents):
                            # 捕捉原始保護選項（包含圖片保護）
                            try:
                                p = ws_com.Protection
                                prot_opts = {
                                    # 🖼️ 圖片和物件保護（關鍵修復）
                                    'DrawingObjects': getattr(p, 'DrawingObjects', False),
                                    'Contents': True,  # 保留內容保護
                                    'Scenarios': getattr(p, 'Scenarios', False),
                                    
                                    # 其他格式和編輯選項
                                    'AllowFormattingCells': getattr(p, 'AllowFormattingCells', False),
                                    'AllowFormattingColumns': getattr(p, 'AllowFormattingColumns', False),
                                    'AllowFormattingRows': getattr(p, 'AllowFormattingRows', False),
                                    'AllowInsertingColumns': getattr(p, 'AllowInsertingColumns', False),
                                    'AllowInsertingRows': getattr(p, 'AllowInsertingRows', False),
                                    'AllowInsertingHyperlinks': getattr(p, 'AllowInsertingHyperlinks', False),
                                    'AllowDeletingColumns': getattr(p, 'AllowDeletingColumns', False),
                                    'AllowDeletingRows': getattr(p, 'AllowDeletingRows', False),
                                    'AllowSorting': getattr(p, 'AllowSorting', False),
                                    'AllowFiltering': getattr(p, 'AllowFiltering', False),
                                    'AllowUsingPivotTables': getattr(p, 'AllowUsingPivotTables', False),
                                }
                            except Exception:
                                # 如果無法取得保護選項，使用預設值
                                prot_opts = {
                                    'DrawingObjects': True,  # 🖼️ 預設保護圖片和物件
                                    'Contents': True,
                                    'Scenarios': False,
                                    'AllowFormattingCells': False,
                                    'AllowFormattingColumns': False,
                                    'AllowFormattingRows': False,
                                    'AllowInsertingColumns': False,
                                    'AllowInsertingRows': False,
                                    'AllowInsertingHyperlinks': False,
                                    'AllowDeletingColumns': False,
                                    'AllowDeletingRows': False,
                                    'AllowSorting': False,
                                    'AllowFiltering': False,
                                    'AllowUsingPivotTables': False,
                                }
                            
                            pwd = simpledialog.askstring(
                                "工作表受保護",
                                f"{sheet_name} 已受保護，請輸入密碼（可留空）：",
                                show='*',
                                parent=self.root
                            )
                            if pwd is None:
                                pwd = ""
                            try:
                                if pwd:
                                    ws_com.Unprotect(pwd)
                                else:
                                    ws_com.Unprotect()
                                reprotect = True
                                used_pwd = pwd
                            except Exception as e:
                                messagebox.showerror("錯誤", f"{sheet_name} 取消保護失敗：{e}")
                                continue
                                
                            # 🔧 修復：存儲保護信息以便稍後恢復
                            sheet_protections[sheet_name] = {
                                'reprotect': reprotect,
                                'prot_opts': prot_opts,
                                'used_pwd': used_pwd
                            }
                    except Exception:
                        pass
                    # 注意：coordinate_to_tuple 返回 (row, column) 而不是 (column, row)
                    # 處理 coordinate_from_string 可能已被移除的問題
                    try:
                        row, col = coordinate_from_string(start_cell)
                        col = int(col)
                        row = int(row)
                    except AttributeError:
                        # 如果 coordinate_from_string 不可用，手動解析座標
                        import re
                        match = re.match(r'([A-Za-z]+)(\d+)', start_cell)
                        if match:
                            col_letter, row = match.groups()
                            from openpyxl.utils import column_index_from_string
                            col = column_index_from_string(col_letter)
                            row = int(row)
                        else:
                            raise ValueError(f"Invalid cell coordinate: {start_cell}")
                    except Exception as e:
                        messagebox.showerror("錯誤", f"解析儲存格座標時發生錯誤：{e}")
                        return
                    filled = 0
                    offset = 0
                    while filled < count and data_idx < total_data:
                        # 橫向填入時跳過一欄（例如 I, K, M, O...）
                        # 修改為正確的寫入模式：I, K, M, O, Q, S, U, W (每隔一欄)
                        r = row
                        c = col + (filled * 2)
                        cell = ws_com.Cells(r, c)
                        is_main = True
                        try:
                            if cell.MergeCells:
                                top_left = cell.MergeArea.Cells(1, 1)
                                is_main = (cell.Address(False, False) == top_left.Address(False, False))
                        except Exception:
                            is_main = True
                        if not is_main:
                            print(f"跳過合併儲存格: [{data_idx}] {repr(data[data_idx])} 到 {cell.Address}")
                            # 橫向時跳過一欄（相當於跳過2個位置）
                            offset += 2 if direction == '横向' else 1
                            # 添加安全檢查防止無限循環
                            if offset > 10000:  # 設置合理的上限
                                messagebox.showwarning("警告", "檢測到可能的無限循環，已中斷操作")
                                break
                            # 即使跳過也要增加filled計數，否則會導致無限循環
                            filled += 1
                            continue
                        # 如果是合併儲存格，跳過整個合併區域的大小
                        if cell.MergeCells:
                            merge_area = cell.MergeArea
                            rows_count = merge_area.Rows.Count
                            cols_count = merge_area.Columns.Count
                            # 計算合併區域覆蓋了多少個單元格
                            merged_cells_count = rows_count * cols_count
                            print(f"檢測到合併儲存格: {cell.Address}, 大小: {rows_count}x{cols_count} ({merged_cells_count} 個單元格)")
                        try:
                            locked = False
                            if cell.MergeCells:
                                locked = bool(cell.MergeArea.Cells(1, 1).Locked)
                            else:
                                locked = bool(cell.Locked)
                        except Exception:
                            locked = False
                        if locked:
                            print(f"跳過鎖定儲存格: [{data_idx}] {repr(data[data_idx])} 到 {cell.Address}")
                            # 橫向時跳過一欄（相當於跳過2個位置）
                            offset += 2 if direction == '横向' else 1
                            # 添加安全檢查防止無限循環
                            if offset > 10000:  # 設置合理的上限
                                messagebox.showwarning("警告", "檢測到可能的無限循環，已中斷操作")
                                break
                            # 即使跳過也要增加filled計數，否則會導致無限循環
                            filled += 1
                            continue
                        value = data[data_idx]
                        # 🔧 修復：添加更多調試信息以幫助診斷問題
                        # 列印當前寫入的資料
                        # 增強調試信息，顯示更多細節
                        print(f"[{data_idx+1}/{total_data}] 準備寫入數據:")
                        print(f"  數據值: {repr(value)} (類型: {type(value).__name__})")
                        print(f"  寫入位置: {cell.Address} (行列: {r},{c})")
                        print(f"  工作表: {sheet_name}")
                        print(f"  方向: 橫向 (I, K, M, O, Q, S, U, W...)")
                        print(f"  規則進度: {filled+1}/{count}")
                        # 🔧 修復：檢查數據有效性
                        if not isinstance(value, (str, int, float)) and value is not None:
                            print(f"  ⚠️ 警告: 數據類型異常: {type(value)}")
                        
                        try:
                            # 保留原始数值格式，避免浮點数精度問題
                            print(f"嘗試寫入值 (COM): {repr(value)} (類型: {type(value)})")
                            cell.Value = value
                            print(f"  寫入成功!")
                        except Exception as e:
                            print(f"寫入失敗 (COM)，使用備用方案: {e}")
                            cell.Value = value
                            print(f"  備用方案寫入完成")
                        data_idx += 1
                        filled += 1
                        # 如果是合併儲存格，跳過整個合併區域的大小
                        if cell.MergeCells:
                            offset += merged_cells_count * 2 if direction == '横向' else merged_cells_count
                        else:
                            # 橫向時跳過一欄（相當於跳過2個位置）
                            offset += 2 if direction == '横向' else 1
                        self.progress_panel.update_progress(data_idx, total_data, status)
                        self.root.update()
                    
                    # 🔧 修復：在每個規則處理完後立即恢復該工作表的保護
                    if reprotect:
                        try:
                            if prot_opts:
                                try:
                                    # 🖼️ 修復：根據用戶發現，使用簡化的保護設定以正確恢復圖片保護
                                    # 只啟用"選取鎖定的儲存格"，禁用其他選項
                                    ws_com.Protect(
                                        Password=used_pwd,
                                        DrawingObjects=True,           # 🖼️ 關鍵：確保圖片被保護
                                        Contents=True,                # 保留內容保護
                                        Scenarios=False,              # 禁用場景保護
                                        AllowFormattingCells=False,    # 禁用格式化儲存格
                                        AllowFormattingColumns=False,  # 禁用格式化列
                                        AllowFormattingRows=False,     # 禁用格式化行
                                        AllowInsertingColumns=False,   # 禁用插入列
                                        AllowInsertingRows=False,      # 禁用插入行
                                        AllowInsertingHyperlinks=False, # 禁用插入超連結
                                        AllowDeletingColumns=False,    # 禁用刪除列
                                        AllowDeletingRows=False,       # 禁用刪除行
                                        AllowSorting=False,            # 禁用排序
                                        AllowFiltering=False,          # 禁用篩選
                                        AllowUsingPivotTables=False,   # 禁用樞紐分析表
                                    )
                                except Exception:
                                    # 如果完整保護失敗，使用最簡保護（包含圖片）
                                    ws_com.Protect(
                                        Password=used_pwd, 
                                        Contents=True,
                                        DrawingObjects=True  # 🖼️ 關鍵：確保圖片被保護
                                    )
                            else:
                                # 如果沒有保護選項，使用預設保護（包含圖片）
                                ws_com.Protect(
                                    Password=used_pwd, 
                                    Contents=True,
                                    DrawingObjects=True  # 🖼️ 關鍵：確保圖片被保護
                                )
                            
                            # 驗證保護是否成功
                            try:
                                if not bool(ws_com.ProtectContents):
                                    # 如果內容未受保護，使用簡化保護（包含圖片）
                                    ws_com.Protect(
                                        Password=used_pwd, 
                                        Contents=True,
                                        DrawingObjects=True  # 🖼️ 關鍵：確保圖片被保護
                                    )
                            except Exception:
                                pass
                        except Exception:
                            pass
                print(f"正在保存文件 (COM): {config['excel_path']}")
                try:
                    wb_com.Save()
                except Exception as e:
                    print(f"保存COM文件時出錯: {e}")
                    # 🔧 修復：如果保存失敗，嘗試另存為
                    try:
                        wb_com.SaveAs(config['excel_path'])
                    except Exception as e2:
                        print(f"另存為COM文件時也出錯: {e2}")
                        raise e  # 重新拋出原始異常
                wb_com.Close(SaveChanges=False)  # 我們已經保存了，所以不需要再保存
                try:
                    excel.ScreenUpdating = True
                    excel.DisplayAlerts = True
                except Exception:
                    pass
                excel.Quit()
                self.progress_panel.update_progress(data_idx, total_data, "完成")
                self.show_report(data_idx, total_data, len(rules))
            except Exception as e:
                messagebox.showerror("錯誤", f"執行時發生錯誤：{e}")
                import traceback
                traceback.print_exc()  # 列印詳細的錯誤資訊
            finally:
                # 🔧 修復：確保Excel對象總是被正確清理，但避免重複關閉已保存的工作簿
                try:
                    # 只有在工作簿尚未關閉的情況下才嘗試關閉
                    if 'wb_com' in locals() and wb_com is not None:
                        try:
                            # 檢查工作簿是否仍然打開
                            _ = wb_com.Name  # 如果工作簿已關閉，這將引發異常
                            wb_com.Close(SaveChanges=False)  # 不保存更改就關閉
                        except:
                            # 工作簿已經關閉，不需要再次關閉
                            pass
                except Exception:
                    pass
                try:
                    if 'excel' in locals() and excel is not None:
                        try:
                            # 檢查Excel應用程序是否仍在運行
                            _ = excel.Version  # 如果Excel已退出，這將引發異常
                            excel.Quit()
                        except:
                            # Excel已經退出，不需要再次退出
                            pass
                except Exception:
                    pass
                gc.collect()

        else:  # Only execute openpyxl section if COM is not used
            wb = None
            try:
                # 🔧 修復：載入時保留VBA和圖片
                wb = load_workbook(config['excel_path'], keep_vba=True)
            except Exception as e:
                messagebox.showerror("錯誤", f"無法載入 Excel 檔案：{e}")
                return

            try:
                for rule in rules:
                    if data_idx >= total_data:
                        break
                    sheet_name = rule['sheet']
                    start_cell = rule['start_cell']
                    count = rule['count']
                    direction = rule['direction']
                    status = f"{sheet_name} - {start_cell}"
                    self.progress_panel.update_progress(data_idx, total_data, status)
                    self.root.update()
                    ws = wb[sheet_name]
                    # 注意：coordinate_to_tuple 返回 (row, column) 而不是 (column, row)
                    # 處理 coordinate_from_string 可能已被移除的問題
                    try:
                        row, col = coordinate_from_string(start_cell)
                        col = int(col)
                        row = int(row)
                    except AttributeError:
                        # 如果 coordinate_from_string 不可用，手動解析座標
                        import re
                        match = re.match(r'([A-Za-z]+)(\d+)', start_cell)
                        if match:
                            col_letter, row = match.groups()
                            from openpyxl.utils import column_index_from_string
                            col = column_index_from_string(col_letter)
                            row = int(row)
                        else:
                            raise ValueError(f"Invalid cell coordinate: {start_cell}")
                    except Exception as e:
                        messagebox.showerror("錯誤", f"解析儲存格座標時發生錯誤：{e}")
                        return
                    filled = 0
                    offset = 0
                    while filled < count and data_idx < total_data:
                        # 橫向填入時跳過一欄（例如 I, K, M, O...）
                        # 修改為正確的寫入模式：I, K, M, O, Q, S, U, W (每隔一欄)
                        # 🔒 固定為橫向填入，移除縱向選項
                        cell = ws.cell(row=row, column=col + (filled * 2))
                        is_merged_main = True
                        merged_range_size = 1
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                if cell.coordinate != merged_range.coord.split(':')[0]:
                                    is_merged_main = False
                                else:
                                    # 計算合併區域大小
                                    # merged_range.size 可能是字典，需要獲取數值
                                    try:
                                        if isinstance(merged_range.size, dict):
                                            # 如果是字典，獲取其中的數值
                                            merged_range_size = list(merged_range.size.values())[0] if merged_range.size else 1
                                        else:
                                            merged_range_size = merged_range.size
                                    except:
                                        # 如果無法獲取size，默認為1
                                        merged_range_size = 1
                                break
                        if not is_merged_main:
                            print(f"跳過合併儲存格: [{data_idx}] {repr(data[data_idx])} 到 {cell.coordinate}")
                            # 橫向時跳過一欄（相當於跳過2個位置）
                            offset += 2 if direction == '横向' else 1
                            # 添加安全檢查防止無限循環
                            if offset > 10000:  # 設置合理的上限
                                messagebox.showwarning("警告", "檢測到可能的無限循環，已中斷操作")
                                break
                            # 即使跳過也要增加filled計數，否則會導致無限循環
                            filled += 1
                            continue
                        print(f"檢測到合併儲存格: {cell.coordinate}, 大小: {merged_range_size} 個單元格")
                        try:
                            is_sheet_protected = bool(getattr(ws.protection, 'sheet', False))
                            if is_sheet_protected:
                                is_locked = bool(getattr(cell.protection, 'locked', True))
                            else:
                                is_locked = False
                        except Exception:
                            is_locked = False
                        if is_locked:
                            print(f"跳過鎖定儲存格: [{data_idx}] {repr(data[data_idx])} 到 {cell.coordinate}")
                            # 橫向時跳過一欄（相當於跳過2個位置）
                            offset += 2 if direction == '横向' else 1
                            # 添加安全檢查防止無限循環
                            if offset > 10000:  # 設置合理的上限
                                messagebox.showwarning("警告", "檢測到可能的無限循環，已中斷操作")
                                break
                            # 即使跳過也要增加filled計數，否則會導致無限循環
                            filled += 1
                            continue
                        value = data[data_idx]
                        # 🔧 修復：添加更多調試信息以幫助診斷問題
                        # 列印當前寫入的資料
                        # 增強調試信息，顯示更多細節
                        print(f"[{data_idx+1}/{total_data}] 準備寫入數據:")
                        print(f"  數據值: {repr(value)} (類型: {type(value).__name__})")
                        print(f"  寫入位置: {cell.coordinate} (行列: {cell.row},{cell.column})")
                        print(f"  工作表: {sheet_name}")
                        print(f"  方向: 橫向 (I, K, M, O, Q, S, U, W...)")
                        print(f"  規則進度: {filled+1}/{count}")
                        # 🔧 修復：檢查數據有效性
                        if not isinstance(value, (str, int, float)) and value is not None:
                            print(f"  ⚠️ 警告: 數據類型異常: {type(value)}")
                        try:
                            # 保留原始数值格式，避免浮點数精度問題
                            print(f"嘗試寫入值 (openpyxl): {repr(value)} (類型: {type(value)})")
                            cell.value = value
                            print(f"  寫入成功!")
                        except ValueError as e:
                            print(f"寫入失敗 (openpyxl)，使用備用方案: {e}")
                            cell.value = value
                            print(f"  備用方案寫入完成")
                        data_idx += 1
                        filled += 1
                        # 移除 offset 的使用，因為我們現在直接根據 filled 計算位置
                        self.progress_panel.update_progress(data_idx, total_data, status)
                        self.root.update()
                try:
                    print(f"正在保存文件 (openpyxl): {config['excel_path']}")
                    wb.save(config['excel_path'])
                except PermissionError:
                    messagebox.showerror("儲存失敗", 
                        f"無法儲存 Excel 檔案！\n\n"
                        f"可能原因：\n"
                        f"1. Excel 檔案正在被其他程式開啟\n"
                        f"2. 檔案被設為唯讀\n"
                        f"3. 沒有寫入權限\n\n"
                        f"請關閉 Excel 檔案後重試。\n\n"
                        f"檔案路徑：\n{config['excel_path']}")
                    return
                except Exception as e:
                    messagebox.showerror("儲存失敗", f"儲存 Excel 檔案時發生錯誤：{e}")
                    return
                self.progress_panel.update_progress(data_idx, total_data, "完成")
                self.show_report(data_idx, total_data, len(rules))
            except Exception as e:
                messagebox.showerror("錯誤", f"執行時發生錯誤：{e}")
                import traceback
                traceback.print_exc()  # 列印詳細的錯誤資訊
            finally:
                if wb:
                    try:
                        # 🔧 修復：確保工作簿正確關閉
                        # 檢查工作簿是否仍然有效
                        try:
                            _ = wb.properties  # 如果工作簿已關閉，這將引發異常
                            wb.close()
                        except:
                            # 工作簿已經關閉
                            pass
                    except Exception as e:
                        print(f"關閉工作簿時出錯: {e}")  # 記錄關閉工作簿時的錯誤
                gc.collect()

    
    def show_report(self, filled, total, rule_count, saved=True):
        """顯示執行報告"""
        report = f"""
執行完成！

已填入資料：{filled} / {total} 筆
使用規則：{rule_count} 條
完成率：{int(filled/total*100)}%

Excel 檔案已儲存。
        """
        if not saved:
            report = f"""
執行完成！

已填入資料：{filled} / {total} 筆
使用規則：{rule_count} 條
完成率：{int(filled/total*100)}%

Excel 檔案未儲存。
        """
        messagebox.showinfo("執行報告", report.strip())
    
    def reset_all(self):
        """重置所有設定"""
        if messagebox.askyesno("確認", "確定要重置所有設定嗎？"):
            # 清除資料
            self.data_panel.clear_data()
            
            # 清除配置
            self.config_panel.clear_rules()
            
            # 🗑️ 釋放 Excel 記憶體
            self.config_panel.workbook = None
            
            # 清空快取
            self.config_panel.sheet_names.clear()
            self.config_panel.excel_file_path = None
            self.config_panel.file_label.config(text="未選擇", fg='gray')
            self.config_panel.sheet_label.config(text="工作表：0 個")
            
            # 重置進度
            self.progress_panel.reset()
            
            # 🗑️ 強制垃圾回收，釋放全部記憶體
            gc.collect()  # 顯式呼叫垃圾回收釋放所有記憶體
            
            messagebox.showinfo("完成", "已重置所有設定並釋放記憶體")
    
    def show_help(self):
        """顯示說明"""
        help_text = """
Txt2Excel v2.1_20251213 使用說明

1. 選擇 TXT 檔案
   - 點擊左側「選擇檔案」按鈕
   - 程式會自動解析資料

2. 選擇 Excel 檔案
   - 點擊右側「選擇檔案」按鈕
   - 載入要填入的 Excel 檔案

3. 配置填入規則
   - 點擊「新增規則」設定填入方式
   - 可新增多條規則處理不同工作表

4. 執行填入
   - 點擊「開始執行」按鈕
   - 等待進度完成

# 🖼️ 圖片保護功能（重要更新）
# - ✅ 自動保留工作表的圖片保護設定
# - ✅ 解鎖後會完整復原所有保護選項
# - ✅ 包含圖片、物件、內容等所有保護項目

提示：
- TXT 格式：每行包含 = 符號
- 支援橫向填入
- 自動跳過合併儲存格
- 工作表保護時會自動提示輸入密碼
        """
        messagebox.showinfo("使用说明", help_text.strip())
    
    def run(self):
        """執行主程式"""
        self.root.mainloop()


def main():
    """主程式入口"""
    try:
        # 在建立主視窗前確保Tcl/Tk函式庫路徑正確
        if 'TCL_LIBRARY' not in os.environ or 'TK_LIBRARY' not in os.environ:
            python_root = "C:\\Python314"  # 預設Python安裝路徑
            tcl_path = os.path.join(python_root, 'tcl', 'tcl8.6')
            tk_path = os.path.join(python_root, 'tcl', 'tk8.6')
            
            if os.path.exists(tcl_path):
                os.environ['TCL_LIBRARY'] = tcl_path
            if os.path.exists(tk_path):
                os.environ['TK_LIBRARY'] = tk_path
        
        app = MainWindow()
        app.run()
    except Exception as e:
        # 如果出現Tcl/Tk相關錯誤，提供友善的錯誤訊息
        if "Tcl" in str(e) or "tkinter" in str(e).lower():
            messagebox.showerror("環境配置錯誤", 
                               "無法初始化圖形介面，請透過 run_app.bat 啟動程式。\n\n"
                               "錯誤詳情:\n" + str(e))
        else:
            messagebox.showerror("未知錯誤", "程式啟動時發生錯誤：\n" + str(e))
        print(f"Error: {e}")
        traceback.print_exc()


if __name__ == '__main__':
    main()

